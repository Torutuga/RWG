from noise import Zone
from time import sleep
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter as _gl
import os, math, random as ran

book = load_workbook('C:/users/hp/dnd/dnd.xlsx')
sheet1 = book['Map1']

def Search(sht, x, y, crit):
    param = [(-1, 0),(-1,-1),(1, 0),(-1,1),(0, -1),(1,1),(1,-1),(0, 1)]
    frontier = []
    group = []

    ori = sht[_gl(x) + str(y)].value

    if ori == crit:
        group.append((x,y))

    for p in param:
        try:
            val = sht[_gl(x+p[0]) + str(y+p[1])].value 
            if val == crit:
                frontier.append((x+p[0], y+p[1]))
        except:
            continue

    while frontier:
        if not frontier:
            break

        data = []
        cr = frontier.pop()

        for p in param:
            try:
                val = sht[_gl(cr[0]+p[0]) + str(cr[1]+p[1])].value 
                if val == crit:
                    data.append((cr[0]+p[0], cr[1]+p[1]))
            except:
                continue

        for x in frontier:
            if x in data:
                data.remove(x)
        for x in group:
            if x in data:
                data.remove(x)

        frontier.extend(data)
        data.clear()
        group.append(cr)

    return group
