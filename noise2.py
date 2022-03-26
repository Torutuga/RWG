from perlin_noise import PerlinNoise
from noise import Roll, Color, Draw, Context, Zone 
from perlinN import noiseGiver
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter as _gl
import os, random as ran, math
from colorama import init

init(autoreset=True)
os.system('cls')
book = load_workbook('C:/Users/Hp/DnD/dnd.xlsx')
sheet1 = book['Map1']

x = input('Ancho? ')
y = input('Alto? ')

x_cor = int(x)
y_cor = int(y)

noiseGiver(sheet1, y_cor, x_cor)
Draw(sheet1, y_cor, x_cor)

#book.save('C:/Users/Hp/DnD/dnd.xlsx')
