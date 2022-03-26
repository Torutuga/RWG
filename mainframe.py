from noise import Roll, Color, Draw, Context, Zone 
from se import Search
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter as _gl
import os, random as ran, math
from colorama import init

init(autoreset=True)
os.system('cls')
book = load_workbook('C:/Users/Hp/DnD/dnd.xlsx')
sheet1 = book['Map1']

answer = input('1. Create\n2.Open\n...')
if answer == '1':
    x = input('Ancho? ')
    y = input('Alto? ')

    x_cor = int(x)
    y_cor = int(y)

    run = True
    while run:
        for y in range(1, y_cor+1):
            for x in range(1, x_cor+1):
                if y == 1 and x == 1:
                    sheet1[_gl(x) + str(y)] = ran.randint(1, 10)
                else:
                    sheet1[_gl(x) + str(y)] = Context(sheet1, x, y)

        for y in range(1, y_cor+1):
            for x in range(1, x_cor+1):
                val = sheet1[_gl(x) + str(y)].value
                if val < 5:
                    result = Context(sheet1, x, y, mode=1)
                    if result:
                        sheet1[_gl(x) + str(y)] = result       

        for y in range(1, y_cor+1):
            for x in range(1, x_cor+1):
                val = sheet1[_gl(x) + str(y)].value
                if val > 4:
                    Context(sheet1, x, y, mode=3)

        water = 0
        land = 0

        for y in range(1, y_cor+1):
            for x in range(1, x_cor+1):
                val = sheet1[_gl(x)+str(y)].value
                if val < 5:
                    water += 1
                if val > 4:
                    land += 1

        if (land*100)/(x_cor*y_cor) >= 60.0:
            run = False
            break
        else:
            for y in range(1, y_cor+1):
                for x in range(1, x_cor+1):
                    del sheet1[_gl(x) + str(y)]

    sheet1['A'+str(y_cor+1)] = str(y_cor) + '/' + str(x_cor)

    Draw(sheet1,y_cor+1,x_cor+1)
    print('\n', water, land)

    reply = input('Save? y/n\n...')
    if reply == 'y':
        book.save('C:/Users/Hp/DnD/dnd.xlsx')

if answer == '2':
    y_cor = 0
    x_cor = 0

    for x in range(1, 100):
        val = sheet1['A'+str(x)].value
        if type(val) == str:
            a, b = val.split('/')
            y_cor = int(a[1:])
            x_cor = int(b)
            break

    Draw(sheet1,y_cor+1,x_cor+1)
            
    pass
